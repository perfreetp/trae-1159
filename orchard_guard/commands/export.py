import os
import csv
import click
from ..core.models import DiseaseType
from ..core.detector import get_treatment
from ..core.store import (
    resolve_sessions, compute_statistics, compute_priority_watch,
    load_config,
)


@click.command("export")
@click.option("--session", "-s", "session_id", default="", help="会话ID")
@click.option("--plot", "-p", default="", help="按地块编号导出")
@click.option("--output", "-o", default="", help="输出文件路径")
@click.option(
    "--format",
    "-f",
    "fmt",
    type=click.Choice(["xlsx", "csv"]),
    default=None,
    help="导出格式",
)
@click.option(
    "--mode",
    "-m",
    type=click.Choice(["detail", "summary"]),
    default="detail",
    help="csv 导出模式: detail=明细, summary=汇总",
)
@click.option("--store-dir", default="", help="数据存储目录")
def export_command(session_id, plot, output, fmt, mode, store_dir):
    """按地块导出表格"""

    config = load_config(store_dir or None)
    export_fmt = fmt or config.export_format

    sessions, ok = resolve_sessions(
        session_id=session_id, plot=plot, store_dir=store_dir or None
    )
    if not ok or not sessions:
        if session_id and plot:
            click.echo(f"❌ 会话 {session_id} 中没有地块 {plot} 的图片")
        elif session_id:
            click.echo(f"❌ 未找到会话: {session_id}")
        elif plot:
            click.echo(f"❌ 地块 {plot} 无可用记录")
        else:
            click.echo("📭 暂无扫描会话")
        return

    if not output:
        if plot:
            output = f"果园病害_{plot}.{export_fmt}"
        elif session_id:
            output = f"果园病害_{session_id}.{export_fmt}"
        else:
            output = f"果园病害_全部.{export_fmt}"

    if export_fmt == "xlsx":
        _export_xlsx(sessions, output, config)
    else:
        if mode == "summary":
            _export_csv_summary(sessions, output, config)
        else:
            _export_csv_detail(sessions, output)

    click.echo(f"✅ 已导出到: {os.path.abspath(output)}")


def _collect_detail_rows(sessions):
    headers = [
        "会话ID", "巡园日期", "品种", "地块编号", "文件名", "文件路径",
        "图片尺寸", "是否模糊", "模糊度", "病害类型", "置信度", "病斑区域",
        "病斑面积(px²)", "面积占比(%)",
        "是否修正", "原始病害",
    ]
    rows = [headers]

    for sess in sessions:
        for img in sess.images:
            img_area = img.image_area()
            size_str = f"{img.image_width}×{img.image_height}" if img_area > 0 else "-"
            if not img.detections:
                rows.append([
                    sess.id, img.scan_date, img.variety, img.plot_id,
                    img.file_name, img.file_path,
                    size_str,
                    "是" if img.is_blurry else "否", f"{img.blur_score:.1f}",
                    "", "", "", "", "", "", "",
                ])
            for det in img.detections:
                bbox_str = ""
                bbox_area = 0
                area_pct = ""
                if det.bbox:
                    bbox_str = f"({det.bbox.x1},{det.bbox.y1})-({det.bbox.x2},{det.bbox.y2})"
                    bbox_area = det.bbox.area()
                    if img_area > 0:
                        area_pct = f"{bbox_area / img_area * 100:.2f}"
                area_str = str(bbox_area) if bbox_area > 0 else ""
                rows.append([
                    sess.id, img.scan_date, img.variety, img.plot_id,
                    img.file_name, img.file_path,
                    size_str,
                    "是" if img.is_blurry else "否", f"{img.blur_score:.1f}",
                    det.disease.value, f"{det.confidence:.1%}",
                    bbox_str,
                    area_str, area_pct,
                    "是" if det.corrected else "否",
                    det.original_disease.value if det.original_disease else "",
                ])
    return rows


def _collect_detail_summary_rows(stats):
    headers = [
        "巡园日期", "地块编号", "品种", "病害", "检出数",
        "病斑面积(px²)", "面积占比(%)",
    ]
    rows = [headers]
    for row in stats.get("detail_summary", []):
        rows.append([
            row["scan_date"], row["plot_id"], row["variety"], row["disease"],
            row["count"], row["lesion_area"], row["area_pct"],
        ])
    return rows


def _collect_plot_summary_rows(stats):
    headers = [
        "地块编号", "总图片数", "病害图片数", "健康图片数", "发病率(%)",
        "病斑面积合计(px²)", "面积占比(%)", "主要病害", "防治建议摘要",
    ]
    rows = [headers]
    plot_stats = stats["plot_stats"]
    for pid in sorted(plot_stats.keys()):
        total_img = stats["plot_total_images"].get(pid, 0)
        disease_img = stats["plot_disease_images"].get(pid, 0)
        healthy_img = stats["plot_healthy_images"].get(pid, 0)
        rate = disease_img / max(1, total_img) * 100
        total_lesion = sum(stats["plot_lesion_area"].get(pid, {}).values())
        total_img_area = stats["plot_image_area"].get(pid, 0)
        area_pct = total_lesion / max(1, total_img_area) * 100 if total_img_area > 0 else 0
        diseases_in_plot = plot_stats[pid]
        if diseases_in_plot:
            primary = max(diseases_in_plot, key=diseases_in_plot.get)
            treatment = get_treatment(primary)
        else:
            primary = "-"
            treatment = "-"
        rows.append([
            pid, total_img, disease_img, healthy_img,
            f"{rate:.1f}", total_lesion, f"{area_pct:.2f}",
            primary, treatment,
        ])
    return rows


def _collect_disease_summary_rows(stats):
    headers = [
        "病害名称", "检出数", "检出占比(%)", "平均置信度",
        "病斑面积合计(px²)", "面积占比(%)", "涉及地块", "防治方案",
    ]
    rows = [headers]
    all_disease_stats = stats["all_disease_stats"]
    plot_stats = stats["plot_stats"]
    total_det = sum(all_disease_stats.values())
    for dname, count in sorted(all_disease_stats.items(), key=lambda x: -x[1]):
        confs = stats["all_confidence"].get(dname, [])
        avg_conf = sum(confs) / len(confs) if confs else 0
        la = stats["disease_lesion_area"].get(dname, 0)
        ia = stats["disease_image_area"].get(dname, 0)
        ratio = count / max(1, total_det) * 100
        area_pct = la / max(1, ia) * 100 if ia > 0 else 0
        affected_plots = [pid for pid in plot_stats if dname in plot_stats[pid]]
        treatment = get_treatment(dname)
        rows.append([
            dname, count, f"{ratio:.1f}", f"{avg_conf:.1%}",
            la, f"{area_pct:.2f}",
            ", ".join(sorted(affected_plots)), treatment,
        ])
    return rows


def _collect_priority_watch_rows(watch):
    headers = [
        "地块编号", "品种", "主要病害", "最近巡园日期",
        "发病率(%)", "面积占比(%)", "增长幅度(%)",
        "触发原因", "防治建议", "建议复查日期",
    ]
    rows = [headers]
    for w in watch:
        if w["triggers"]:
            rows.append([
                w["plot_id"], w["variety"], w["primary_disease"], w["latest_date"],
                w["incidence_rate"], w["area_pct"], w["growth"],
                "; ".join(w["triggers"]), w["treatment"], w["recheck_date"],
            ])
    return rows


def _export_csv_detail(sessions, output_path):
    rows = _collect_detail_rows(sessions)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)
    click.echo(f"   明细: {len(rows)-1} 条记录")


def _export_csv_summary(sessions, output_path, config):
    stats = compute_statistics(sessions)
    watch = compute_priority_watch(sessions, config)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["=== 巡园台账汇总 ==="])
        for row in _collect_detail_summary_rows(stats):
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(["=== 地块台账汇总 ==="])
        for row in _collect_plot_summary_rows(stats):
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(["=== 病害汇总 ==="])
        for row in _collect_disease_summary_rows(stats):
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(["=== 重点巡查清单 ==="])
        for row in _collect_priority_watch_rows(watch):
            writer.writerow(row)
    click.echo(f"   汇总: 台账明细 + 地块汇总 + 病害汇总 + 重点清单")


def _export_xlsx(sessions, output_path, config):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        click.echo("⚠️  未安装 openpyxl，改用 CSV 格式导出")
        csv_path = output_path.rsplit(".", 1)[0] + ".csv"
        _export_csv_detail(sessions, csv_path)
        return

    stats = compute_statistics(sessions)
    watch = compute_priority_watch(sessions, config)
    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    disease_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alert_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws_detail = wb.active
    ws_detail.title = "明细"
    _write_sheet(ws_detail, _collect_detail_rows(sessions), header_fill, header_font, thin_border, disease_fill)

    ws_ledger = wb.create_sheet("巡园台账汇总")
    _write_sheet(ws_ledger, _collect_detail_summary_rows(stats), header_fill, header_font, thin_border, disease_fill)

    ws_plot = wb.create_sheet("地块汇总")
    _write_sheet(ws_plot, _collect_plot_summary_rows(stats), header_fill, header_font, thin_border)

    ws_disease = wb.create_sheet("病害汇总")
    _write_sheet(ws_disease, _collect_disease_summary_rows(stats), header_fill, header_font, thin_border, disease_fill)

    ws_watch = wb.create_sheet("重点巡查清单")
    watch_rows = _collect_priority_watch_rows(watch)
    _write_sheet(ws_watch, watch_rows, header_fill, header_font, thin_border, disease_fill, alert_fill)

    wb.save(output_path)
    click.echo(f"   5张表: 明细 + 巡园台账汇总 + 地块汇总 + 病害汇总 + 重点巡查清单")


def _write_sheet(ws, rows, header_fill, header_font, border, disease_fill=None, alert_fill=None):
    from openpyxl.styles import Alignment as _Alignment

    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = _Alignment(vertical="center")
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    if disease_fill or alert_fill:
        for row_idx in range(2, len(rows) + 1):
            for col_idx in range(1, len(rows[0]) + 1):
                header = rows[0][col_idx - 1] if col_idx <= len(rows[0]) else ""
                val = ws.cell(row=row_idx, column=col_idx).value
                if not val:
                    continue
                if header in ("病害类型", "病害名称", "主要病害", "病害") and val not in ("健康", "", "-"):
                    ws.cell(row=row_idx, column=col_idx).fill = disease_fill or alert_fill
                if header == "触发原因" and val:
                    ws.cell(row=row_idx, column=col_idx).fill = alert_fill or disease_fill

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)
